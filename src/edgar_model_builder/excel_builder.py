from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def _style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="F2F2F2")
    for cell in ws[row]:
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = fill

def _autowidth(ws):
    for col in ws.columns:
        maxlen = 0
        for c in col:
            v = "" if c.value is None else str(c.value)
            maxlen = max(maxlen, len(v))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(maxlen + 2, 44)

def _write_df(ws, df: pd.DataFrame, index_name: str = "end"):
    ws.append([index_name] + list(df.columns))
    for idx, row in df.sort_index().iterrows():
        ws.append([idx.date().isoformat()] + [None if pd.isna(row[c]) else float(row[c]) for c in df.columns])
    _style_header(ws, 1)
    ws.freeze_panes = "A2"
    _autowidth(ws)

def build_model_xlsx(out_path: str | Path, target_ticker: str, hist_fy: pd.DataFrame, hist_q: pd.DataFrame, comps: pd.DataFrame, assumptions: dict):
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Summary")
    ws["A1"] = "Ticker"
    ws["B1"] = target_ticker
    ws["A3"] = "Assumption"
    ws["B3"] = "Value"
    r = 4
    for k, v in assumptions.items():
        ws[f"A{r}"] = k
        ws[f"B{r}"] = v
        r += 1
    ws["A10"] = "Notes"
    ws["B10"] = "Update assumptions + LBO inputs, then refresh outputs."
    _style_header(ws, 3)
    _autowidth(ws)

    ws = wb.create_sheet("Hist_FY")
    _write_df(ws, hist_fy, "end")

    ws = wb.create_sheet("Hist_Q")
    _write_df(ws, hist_q, "end")

    ws = wb.create_sheet("Comps")
    ws.append(list(comps.columns))
    for _, row in comps.iterrows():
        ws.append([None if pd.isna(row[c]) else row[c] for c in comps.columns])
    _style_header(ws, 1)
    ws.freeze_panes = "A2"
    _autowidth(ws)

    ws = wb.create_sheet("LBO")
    ws["A1"] = "LBO Module (Starter)"
    ws["A3"] = "Input"
    ws["B3"] = "Value"
    inputs = [
        ("TTM EBITDA", None),
        ("Entry EV/EBITDA", 10.0),
        ("Debt / EBITDA", 5.0),
        ("Interest Rate", 0.085),
        ("Holding Period (Years)", 5),
        ("Exit EV/EBITDA", 10.0),
        ("EBITDA CAGR", 0.06),
        ("Annual Debt Paydown (% of Debt)", 0.10),
    ]
    for i, (k, v) in enumerate(inputs, start=4):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v
    _style_header(ws, 3)
    ws["D3"] = "Outputs"
    ws["E3"] = "Value"
    _style_header(ws, 3)

    ws["D4"] = "Entry EV"
    ws["D5"] = "Entry Debt"
    ws["D6"] = "Entry Equity"
    ws["D7"] = "Exit EBITDA"
    ws["D8"] = "Exit EV"
    ws["D9"] = "Exit Debt"
    ws["D10"] = "Exit Equity"
    ws["D11"] = "MOIC"
    ws["D12"] = "Equity IRR"

    ws["B4"] = '=IFERROR(INDEX(Comps!$G:$G, MATCH("' + target_ticker + '", Comps!$A:$A, 0)), "")'
    ws["B4"].number_format = "0.00"
    ws["B5"].number_format = "0.00"
    ws["B6"].number_format = "0.00"
    ws["B7"].number_format = "0.00%"
    ws["B8"].number_format = "0"
    ws["B9"].number_format = "0.00"
    ws["B10"].number_format = "0.00"
    ws["B11"].number_format = "0.00"
    ws["B12"].number_format = "0.00"
    ws["B13"].number_format = "0.00%"
    ws["B14"].number_format = "0.00%"

    ws["E4"] = "=B4*B5"
    ws["E5"] = "=B4*B6"
    ws["E6"] = "=E4-E5"
    ws["E7"] = "=B4*(1+B10)^B8"
    ws["E8"] = "=E7*B9"
    ws["E9"] = "=E5*(1-B11)^B8"
    ws["E10"] = "=E8-E9"
    ws["E11"] = "=E10/E6"
    ws["E12"] = '=IFERROR((E11)^(1/B8)-1, "")'
    for rr in range(4, 12):
        ws[f"E{rr}"].number_format = "0.00"
    ws["E11"].number_format = "0.00x"
    ws["E12"].number_format = "0.00%"
    ws.freeze_panes = "A4"
    _autowidth(ws)

    wb.save(out_path)
    return str(out_path)
